from PyQt5 import QtCore, QtGui, QtWidgets
from Class_filtr import *
from Class_for_request_directory import *
import os
from openpyxl import Workbook
import sqlite3
import csv


class class_availability_of_wagons_widget(
    QtWidgets.QWidget, stack_level_1, stack_level_2
):
    def __init__(self, table):
        super().__init__()
        self.name = "Наличие вагонов"

        self.grid = QtWidgets.QGridLayout()

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")
        self.grid.addWidget(self.text_from, 1, 1)

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))
        self.grid.addWidget(self.date_from, 1, 2)

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")
        self.grid.addWidget(self.text_to, 1, 3)

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))
        self.grid.addWidget(self.date_to, 1, 4)

        self.number_vagon = QtWidgets.QLabel()
        self.number_vagon.setText("№ вагона")
        self.grid.addWidget(self.number_vagon, 1, 5)

        self.line = QtWidgets.QLineEdit(self)
        self.grid.addWidget(self.line, 1, 6)

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))
        self.btn_update.clicked.connect(self.write)
        self.grid.addWidget(self.btn_update, 1, 7)

        self.to_excel = QtWidgets.QPushButton("", self)
        self.to_excel.setIcon(QtGui.QIcon("excel.png"))
        self.grid.addWidget(self.to_excel, 1, 8)

        self.table = table

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)


class class_availability_of_wagons(
    QtWidgets.QWidget, class_filtr, class_for_request_directory
):
    def __init__(self):
        super().__init__()

        self.btn_update = QtWidgets.QPushButton("Обновить все", self)
        self.btn_update.clicked.connect(self.write)

        self.to_excel = QtWidgets.QPushButton("В Excel все", self)
        self.to_excel.clicked.connect(self.write_excel)

        self.dict_class = {"Наличие вагонов": class_availability_of_wagons_widget}

        self.dict_table = {"Наличие вагонов": None}

        self.grid = QtWidgets.QGridLayout()
        self.write_table()

        self.grid.addWidget(self.btn_update, 1, 1)
        self.grid.addWidget(self.to_excel, 1, 2)

        self.setLayout(self.grid)

    def write_table(self):
        for i in self.dict_table:
            self.dict_table[i] = QtWidgets.QTableWidget(self)
            self.dict_table[i].setColumnCount(0)
            self.dict_table[i].setRowCount(0)

            self.table_of_contents = self.dict_table[i].horizontalHeader()
            self.table_of_contents.sectionClicked.connect(self.click_table_of_contents)

            item = self.dict_class[i](self.dict_table[i])

            self.grid.addWidget(item, 2, 1, 1, 14)

    def write(self):
        self.dict_list = {
            "Наличие вагонов": list(
                csv.reader(open("Наличие вагонов.csv", "r"), delimiter=";")
            )
        }

        for i in self.dict_list:
            self.write_list(i, self.dict_list[i])

    def write_excel(self):
        os.chdir(
            str(
                QtWidgets.QFileDialog.getExistingDirectory(
                    self, "Выберите папку для сохранения"
                )
            )
        )
        wb = Workbook()

        for i in self.dict_list:
            wb.create_sheet(i)
            sheet = wb[i]
            for row in range(len(self.dict_list[i])):

                shiht_row = row + 1

                for column in range(len(self.dict_list[i][row])):
                    sheet.cell(row=shiht_row, column=column + 1).value = self.dict_list[
                        i
                    ][row][column]

            sheet.auto_filter.ref = "A1:AK100000"
            sheet.freeze_panes = sheet["A2"]
            sheet.sheet_view.zoomScale = 70
        wb.remove(wb["Sheet"])

        wb.save("excel.xlsx")
        print("excel.xlsx - done...")
