from PyQt5 import QtCore, QtGui, QtWidgets
from Class_filtr import *
from Class_for_request_directory import *
import os
from openpyxl import Workbook
import sqlite3
import csv


class Class_an_approach_widget(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Подход"

        self.btn_update = QtWidgets.QPushButton("Обновить", self)
        self.btn_update.clicked.connect(self.write)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))

        self.btn_load_gvc = QtWidgets.QPushButton("Загрузка гвц", self)
        self.btn_load_gvc.clicked.connect(self.load_gvc_on_click)
        self.btn_load_gvc.setIcon(QtGui.QIcon("load.png"))

        self.btn_connect_cargo = QtWidgets.QPushButton("Связать груз", self)
        self.btn_connect_cargo.setIcon(QtGui.QIcon("connect cargo.png"))

        self.btn_connect_firm = QtWidgets.QPushButton("Связать груз", self)
        self.btn_connect_firm.setIcon(QtGui.QIcon("connect firm.png"))

        self.btn_make = QtWidgets.QPushButton("Связать груз", self)
        self.btn_make.setIcon(QtGui.QIcon("train.png"))

        self.table = table

        self.grid = QtWidgets.QGridLayout()
        self.grid.addWidget(self.btn_update, 1, 1)
        self.grid.addWidget(self.btn_load_gvc, 1, 2)
        self.grid.addWidget(self.btn_connect_cargo, 1, 3)
        self.grid.addWidget(self.btn_connect_firm, 1, 4)
        self.grid.addWidget(self.btn_make, 1, 5)

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)

    def load_gvc_on_click(self):
        file = QtWidgets.QFileDialog.getOpenFileName()
        with open(file[0]) as f:
            lines = f.readlines()
            print(lines)
        with open(self.name + ".csv", "a") as file_object:
            for line in lines:
                file_object.write(line)
        self.write()


class Class_admission_widget(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Поступление"

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))

        self.btn_find_by_period = QtWidgets.QPushButton("", self)
        self.btn_find_by_period.clicked.connect(self.write)
        self.btn_find_by_period.setIcon(QtGui.QIcon("green check.png"))

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))

        self.btn_make_record = QtWidgets.QPushButton("", self)
        self.btn_make_record.setIcon(QtGui.QIcon("make record.png"))

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))

        self.btn_delete_record = QtWidgets.QPushButton("", self)
        self.btn_delete_record.setIcon(QtGui.QIcon("delete record.png"))

        self.btn_report = QtWidgets.QPushButton("", self)
        self.btn_report.setIcon(QtGui.QIcon("clock.png"))

        self.grid = QtWidgets.QGridLayout()
        self.table = table

        self.grid.addWidget(self.text_from, 1, 1)
        self.grid.addWidget(self.date_from, 1, 2)
        self.grid.addWidget(self.text_to, 1, 3)
        self.grid.addWidget(self.date_to, 1, 4)
        self.grid.addWidget(self.btn_find_by_period, 1, 5)
        self.grid.addWidget(self.btn_update, 1, 6)
        self.grid.addWidget(self.btn_make_record, 1, 7)
        self.grid.addWidget(self.btn_edit_record, 1, 8)
        self.grid.addWidget(self.btn_delete_record, 1, 9)
        self.grid.addWidget(self.btn_report, 1, 10)

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)


class Class_arrival_notifications(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Уведомления о прибытии"

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))

        self.btn_find_by_period = QtWidgets.QPushButton("", self)
        self.btn_find_by_period.clicked.connect(self.write)
        self.btn_find_by_period.setIcon(QtGui.QIcon("green check.png"))

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))

        self.btn_make_record = QtWidgets.QPushButton("", self)
        self.btn_make_record.setIcon(QtGui.QIcon("make record.png"))

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))

        self.btn_delete_record = QtWidgets.QPushButton("", self)
        self.btn_delete_record.setIcon(QtGui.QIcon("delete record.png"))

        self.grid = QtWidgets.QGridLayout()
        self.table = table

        self.grid.addWidget(self.text_from, 1, 1)
        self.grid.addWidget(self.date_from, 1, 2)
        self.grid.addWidget(self.text_to, 1, 3)
        self.grid.addWidget(self.date_to, 1, 4)
        self.grid.addWidget(self.btn_find_by_period, 1, 5)
        self.grid.addWidget(self.btn_update, 1, 6)
        self.grid.addWidget(self.btn_make_record, 1, 7)
        self.grid.addWidget(self.btn_edit_record, 1, 8)
        self.grid.addWidget(self.btn_delete_record, 1, 9)

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)


class Class_park_edit_moving_window(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Движение вагонов")
        self.setWindowIcon(QtGui.QIcon("train.png"))
        self.setGeometry(0, 0, 200, 100)

        self.grid = QtWidgets.QGridLayout()

        self.select = QtWidgets.QComboBox()
        self.select.addItems(["K01", "K02"])
        self.grid.addWidget(self.select, 1, 1)

        self.int_input = QtWidgets.QSpinBox(self)
        self.grid.addWidget(self.int_input, 1, 2)

        self.text = QtWidgets.QLabel()
        self.text.setText("Комментарий:")
        self.grid.addWidget(self.text, 2, 1, 2, 2)

        self.line = QtWidgets.QLineEdit(self)
        self.grid.addWidget(self.line, 3, 1, 3, 2)

        self.setLayout(self.grid)

        coordinate = self.frameGeometry()
        coordinate.moveCenter(QtWidgets.QDesktopWidget().availableGeometry().center())
        self.move(coordinate.topLeft())


class Class_park_moving_window(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Перемещение вагонов")
        self.setWindowIcon(QtGui.QIcon("train.png"))
        self.setFixedSize(650, 500)

        self.grid = QtWidgets.QGridLayout()
        self.grid.columnMinimumWidth(3)

        self.select = QtWidgets.QComboBox()
        self.select.addItems(["K01", "K02"])
        self.grid.addWidget(self.select, 0, 0, 1, 0)

        self.table_park_moving_window_1 = QtWidgets.QTableWidget()
        self.table_park_moving_window_1.setColumnCount(5)
        self.table_park_moving_window_1.setRowCount(10)
        self.table_park_moving_window_1.setHorizontalHeaderLabels(
            ["№ Вагона", "№ ЖД", "№ Вагона", "Последнее перемещение", "Комментарий"]
        )
        table_items_1 = [
            [
                "59885798",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59881904",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "52469442",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "52469434",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59887827",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59884833",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59940007",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
        ]

        for row in range(1, len(table_items_1)):
            for column in range(len(table_items_1[0])):
                self.table_park_moving_window_1.setItem(
                    row - 1,
                    column,
                    QtWidgets.QTableWidgetItem(table_items_1[row][column]),
                )

        self.grid.addWidget(self.table_park_moving_window_1, 1, 0, 15, 1)

        self.btn_left = QtWidgets.QPushButton("<", self)
        self.grid.addWidget(self.btn_left, 1, 2, 1, 3)
        self.btn_right = QtWidgets.QPushButton(">", self)
        self.grid.addWidget(self.btn_right, 2, 2, 2, 3)

        self.btn_left_2 = QtWidgets.QPushButton("<<", self)
        self.grid.addWidget(self.btn_left_2, 3, 2, 3, 3)
        self.btn_right_2 = QtWidgets.QPushButton(">>", self)
        self.grid.addWidget(self.btn_right_2, 4, 2, 4, 3)
        self.btn_close = QtWidgets.QPushButton("Закрыть", self)
        self.grid.addWidget(self.btn_close, 5, 2, 5, 3)
        ##

        self.table_park_moving_window_2 = QtWidgets.QTableWidget()
        self.table_park_moving_window_2.setColumnCount(5)
        self.table_park_moving_window_2.setRowCount(10)
        self.table_park_moving_window_2.setHorizontalHeaderLabels(
            ["№ Вагона", "№ ЖД", "№ Вагона", "Последнее перемещение", "Комментарий"]
        )

        table_items_2 = [
            [
                "59885798",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59881904",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "52469442",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "52469434",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59887827",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59884833",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
            [
                "59940007",
                "К01",
                "2",
                "14.10.10 04:52",
                "",
            ],
        ]

        for row in range(1, len(table_items_2)):
            for column in range(len(table_items_2[0])):
                self.table_park_moving_window_2.setItem(
                    row - 1,
                    column,
                    QtWidgets.QTableWidgetItem(table_items_2[row][column]),
                )

        self.grid.addWidget(self.table_park_moving_window_2, 1, 6, 15, 9)

        self.setLayout(self.grid)

        coordinate = self.frameGeometry()
        coordinate.moveCenter(QtWidgets.QDesktopWidget().availableGeometry().center())
        self.move(coordinate.topLeft())


class Class_park(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Парк"
        self.grid = QtWidgets.QGridLayout()

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))
        self.btn_edit_record.clicked.connect(self.edit_moving_window_event)
        self.grid.addWidget(self.btn_edit_record, 1, 1)

        self.btn_window_moving = QtWidgets.QPushButton("", self)
        self.btn_window_moving.setIcon(QtGui.QIcon("arrows.png"))
        self.btn_window_moving.clicked.connect(self.moving_window_event)
        self.grid.addWidget(self.btn_window_moving, 1, 2)

        self.btn_make_list = QtWidgets.QPushButton("", self)
        self.btn_make_list.setIcon(QtGui.QIcon("train.png"))
        self.grid.addWidget(self.btn_make_list, 1, 3)

        self.btn_to_front_load = QtWidgets.QPushButton("", self)
        self.btn_to_front_load.setIcon(QtGui.QIcon("warehouse.png"))
        self.grid.addWidget(self.btn_to_front_load, 1, 4)

        self.btn_maneuver_list = QtWidgets.QPushButton("", self)
        self.btn_maneuver_list.setIcon(QtGui.QIcon("green car.png"))
        self.grid.addWidget(self.btn_maneuver_list, 1, 5)

        self.btn_make_mark = QtWidgets.QPushButton("", self)
        self.btn_make_mark.setIcon(QtGui.QIcon("green circle.png"))
        self.grid.addWidget(self.btn_make_mark, 1, 6)

        self.btn_cancel_mark = QtWidgets.QPushButton("", self)
        self.btn_cancel_mark.setIcon(QtGui.QIcon("cancel.png"))
        self.grid.addWidget(self.btn_cancel_mark, 1, 7)

        self.btn_make_comment = QtWidgets.QPushButton("", self)
        self.btn_make_comment.setIcon(QtGui.QIcon("blue i.png"))
        self.grid.addWidget(self.btn_make_comment, 1, 8)

        self.btn_change_order = QtWidgets.QPushButton("", self)
        self.btn_change_order.setIcon(QtGui.QIcon("arrows around.png"))
        self.grid.addWidget(self.btn_change_order, 1, 9)

        self.btn_order_numbers = QtWidgets.QPushButton("", self)
        self.btn_order_numbers.setIcon(QtGui.QIcon("arrow down.png"))
        self.grid.addWidget(self.btn_order_numbers, 1, 10)

        self.btn_order_by_numbers = QtWidgets.QPushButton("", self)
        self.btn_order_by_numbers.setIcon(QtGui.QIcon("arrows 1 to 2.png"))
        self.grid.addWidget(self.btn_order_by_numbers, 1, 11)

        self.btn_update_from_1c = QtWidgets.QPushButton("", self)
        self.btn_update_from_1c.setIcon(QtGui.QIcon("to dir.png"))
        self.grid.addWidget(self.btn_update_from_1c, 1, 12)

        self.btn_update_status = QtWidgets.QPushButton("", self)
        self.btn_update_status.setIcon(QtGui.QIcon("look.png"))
        self.grid.addWidget(self.btn_update_status, 1, 13)

        self.btn_show_all_cars = QtWidgets.QPushButton("", self)
        self.btn_show_all_cars.setIcon(QtGui.QIcon("star.png"))
        self.grid.addWidget(self.btn_show_all_cars, 1, 14)

        self.btn_to_excel = QtWidgets.QPushButton("", self)
        self.btn_to_excel.setIcon(QtGui.QIcon("excel.png"))
        self.grid.addWidget(self.btn_to_excel, 1, 15)

        self.text = QtWidgets.QLabel()
        self.text.setText("Местополож.:")
        self.grid.addWidget(self.text, 1, 16)

        self.select = QtWidgets.QComboBox()
        self.select.addItems(["Неназначеные"])
        self.grid.addWidget(self.select, 1, 17)

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))
        self.btn_update.clicked.connect(self.write)
        self.grid.addWidget(self.btn_update, 1, 18)

        self.select2 = QtWidgets.QComboBox()
        self.select2.addItems(["№ ЖД пути"])
        self.grid.addWidget(self.select2, 2, 1)

        self.table = table
        self.grid.addWidget(self.table, 3, 1, 2, 30)

        self.setLayout(self.grid)

    def edit_moving_window_event(self):
        self.edit_moving_window = Class_park_edit_moving_window()
        self.edit_moving_window.show()

    def moving_window_event(self):
        self.edit_moving_window = Class_park_moving_window()
        self.edit_moving_window.show()


class Class_shunting_full_scale_sheets(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Маневровые натурные листы"

        self.grid = QtWidgets.QGridLayout()

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")
        self.grid.addWidget(self.text_from, 1, 1)

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))
        self.grid.addWidget(self.date_from, 1, 2)

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")
        self.grid.addWidget(self.text_to, 1, 3)

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))
        self.grid.addWidget(self.date_to, 1, 4)

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))
        self.btn_update.clicked.connect(self.write)
        self.grid.addWidget(self.btn_update, 1, 6)

        self.btn_make_record = QtWidgets.QPushButton("", self)
        self.btn_make_record.setIcon(QtGui.QIcon("make record.png"))
        self.grid.addWidget(self.btn_make_record, 1, 7)

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))
        self.grid.addWidget(self.btn_edit_record, 1, 8)

        self.btn_delete = QtWidgets.QPushButton("", self)
        self.btn_delete.setIcon(QtGui.QIcon("red circle.png"))
        self.grid.addWidget(self.btn_delete, 1, 9)

        self.table = table

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)


class Class_load_operations(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Грузовые операции"

        self.grid = QtWidgets.QGridLayout()

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")
        self.grid.addWidget(self.text_from, 1, 1)

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))
        self.grid.addWidget(self.date_from, 1, 2)

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")
        self.grid.addWidget(self.text_to, 1, 3)

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))
        self.grid.addWidget(self.date_to, 1, 4)

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))
        self.btn_update.clicked.connect(self.write)
        self.grid.addWidget(self.btn_update, 1, 5)

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))
        self.grid.addWidget(self.btn_edit_record, 1, 6)

        self.btn_delete = QtWidgets.QPushButton("", self)
        self.btn_delete.setIcon(QtGui.QIcon("red circle.png"))
        self.grid.addWidget(self.btn_delete, 1, 7)

        self.table = table

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)


class Class_departure(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Отправление"

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))

        self.btn_find_by_period = QtWidgets.QPushButton("", self)
        self.btn_find_by_period.clicked.connect(self.write)
        self.btn_find_by_period.setIcon(QtGui.QIcon("green check.png"))

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))

        self.btn_make_record = QtWidgets.QPushButton("", self)
        self.btn_make_record.setIcon(QtGui.QIcon("make record.png"))

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))

        self.btn_delete_record = QtWidgets.QPushButton("", self)
        self.btn_delete_record.setIcon(QtGui.QIcon("delete record.png"))

        self.btn_report = QtWidgets.QPushButton("", self)
        self.btn_report.setIcon(QtGui.QIcon("clock.png"))

        self.grid = QtWidgets.QGridLayout()
        self.table = table

        self.grid.addWidget(self.text_from, 1, 1)
        self.grid.addWidget(self.date_from, 1, 2)
        self.grid.addWidget(self.text_to, 1, 3)
        self.grid.addWidget(self.date_to, 1, 4)
        self.grid.addWidget(self.btn_find_by_period, 1, 5)
        self.grid.addWidget(self.btn_update, 1, 6)
        self.grid.addWidget(self.btn_make_record, 1, 7)
        self.grid.addWidget(self.btn_edit_record, 1, 8)
        self.grid.addWidget(self.btn_delete_record, 1, 9)
        self.grid.addWidget(self.btn_report, 1, 10)

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)


class Class_surrender_notices(QtWidgets.QWidget, stack_level_1, stack_level_2):
    def __init__(self, table):
        super().__init__()
        self.name = "Уведомления о сдаче"

        self.grid = QtWidgets.QGridLayout()

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")
        self.grid.addWidget(self.text_from, 1, 1)

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))
        self.grid.addWidget(self.date_from, 1, 2)

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")
        self.grid.addWidget(self.text_to, 1, 3)

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))
        self.grid.addWidget(self.date_to, 1, 4)

        self.btn_find_by_period = QtWidgets.QPushButton("", self)
        self.btn_find_by_period.setIcon(QtGui.QIcon("green check.png"))
        self.grid.addWidget(self.btn_find_by_period, 1, 5)

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))
        self.btn_update.clicked.connect(self.write)
        self.grid.addWidget(self.btn_update, 1, 5)

        self.btn_make_record = QtWidgets.QPushButton("", self)
        self.btn_make_record.setIcon(QtGui.QIcon("make record.png"))
        self.grid.addWidget(self.btn_make_record, 1, 6)

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))
        self.grid.addWidget(self.btn_edit_record, 1, 7)

        self.btn_delete = QtWidgets.QPushButton("", self)
        self.btn_delete.setIcon(QtGui.QIcon("red circle.png"))
        self.grid.addWidget(self.btn_delete, 1, 8)

        self.btn_notepad = QtWidgets.QPushButton("", self)
        self.btn_notepad.setIcon(QtGui.QIcon("notepad.png"))
        self.grid.addWidget(self.btn_notepad, 1, 9)

        self.table = table

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)


class class_an_approach(QtWidgets.QWidget, class_filtr, class_for_request_directory):
    def __init__(self):
        super().__init__()

        self.btn_update = QtWidgets.QPushButton("Обновить все", self)
        self.btn_update.clicked.connect(self.write)

        self.to_excel = QtWidgets.QPushButton("В Excel все", self)
        self.to_excel.clicked.connect(self.write_excel)

        self.dict_class = {
            "Подход": Class_an_approach_widget,
            "Поступление": Class_admission_widget,
            "Уведомления о прибытии": Class_arrival_notifications,
            "Парк": Class_park,
            "Маневровые натурные листы": Class_shunting_full_scale_sheets,
            "Грузовые операции": Class_load_operations,
            "Отправление": Class_departure,
            "Уведомления о сдаче": Class_surrender_notices,
        }

        self.dict_table = {
            "Подход": None,
            "Поступление": None,
            "Уведомления о прибытии": None,
            "Парк": None,
            "Маневровые натурные листы": None,
            "Грузовые операции": None,
            "Отправление": None,
            "Уведомления о сдаче": None,
        }
        self.tab = QtWidgets.QTabWidget()
        self.write_table()

        self.grid = QtWidgets.QGridLayout()

        self.grid.addWidget(self.btn_update, 1, 1)
        self.grid.addWidget(self.to_excel, 1, 2)
        self.grid.addWidget(self.tab, 2, 1, 1, 14)

        self.setLayout(self.grid)

    def write(self):
        self.dict_list = {
            "Подход": list(csv.reader(open("Подход.csv", "r"), delimiter=";")),
            "Поступление": list(
                csv.reader(open("Поступление.csv", "r"), delimiter=";")
            ),
            "Уведомления о прибытии": list(
                csv.reader(open("Уведомления о прибытии.csv", "r"), delimiter=";")
            ),
            "Парк": list(csv.reader(open("Парк.csv", "r"), delimiter=";")),
            "Маневровые натурные листы": list(
                csv.reader(open("Маневровые натурные листы.csv", "r"), delimiter=";")
            ),
            "Грузовые операции": list(
                csv.reader(open("Грузовые операции.csv", "r"), delimiter=";")
            ),
            "Отправление": list(
                csv.reader(open("Отправление.csv", "r"), delimiter=";")
            ),
            "Уведомления о сдаче": list(
                csv.reader(open("Уведомления о сдаче.csv", "r"), delimiter=";")
            ),
        }

        for i in self.dict_list:
            self.write_list(i, self.dict_list[i])

    def write_excel(self):
        os.chdir(
            str(
                QtWidgets.QFileDialog.getExistingDirectory(
                    self, "Выберите папку для сохранения"
                )
            )
        )
        wb = Workbook()

        for i in self.dict_list:
            wb.create_sheet(i)
            sheet = wb[i]
            for row in range(len(self.dict_list[i])):

                shiht_row = row + 1

                for column in range(len(self.dict_list[i][row])):
                    sheet.cell(row=shiht_row, column=column + 1).value = self.dict_list[
                        i
                    ][row][column]

            sheet.auto_filter.ref = "A1:AK100000"
            sheet.freeze_panes = sheet["A2"]
            sheet.sheet_view.zoomScale = 70
        wb.remove(wb["Sheet"])

        wb.save("excel.xlsx")
        print("excel.xlsx - done...")
