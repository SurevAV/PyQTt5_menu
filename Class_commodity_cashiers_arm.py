from PyQt5 import QtCore, QtGui, QtWidgets
from Class_filtr import *
from Class_for_request_directory import *
import os
from openpyxl import Workbook
import sqlite3
import csv


class Class_commodity_cashiers_arm_window(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Редактирование информации по вагону")
        self.setWindowIcon(QtGui.QIcon("train.png"))

        self.label = QtWidgets.QLabel(self)

        self.pixmap = QtGui.QPixmap("image.png")

        self.label.setPixmap(self.pixmap)

        self.label.resize(self.pixmap.width(), self.pixmap.height())
        self.setGeometry(0, 0, self.pixmap.width(), self.pixmap.height())

        coordinate = self.frameGeometry()
        coordinate.moveCenter(QtWidgets.QDesktopWidget().availableGeometry().center())
        self.move(coordinate.topLeft())


class class_commodity_cashiers_arm_widget(
    QtWidgets.QWidget, stack_level_1, stack_level_2
):
    def __init__(self, table):
        super().__init__()
        self.name = "АРМ товарного кассира"

        self.grid = QtWidgets.QGridLayout()

        self.text_from = QtWidgets.QLabel()
        self.text_from.setText("C")
        self.grid.addWidget(self.text_from, 1, 1)

        self.date_from = QtWidgets.QDateEdit()
        self.date_from.setDate(QtCore.QDate(2022, 6, 10))
        self.grid.addWidget(self.date_from, 1, 2)

        self.text_to = QtWidgets.QLabel()
        self.text_to.setText("по")
        self.grid.addWidget(self.text_to, 1, 3)

        self.date_to = QtWidgets.QDateEdit()
        self.date_to.setDate(QtCore.QDate(2022, 10, 12))
        self.grid.addWidget(self.date_to, 1, 4)

        self.line = QtWidgets.QLineEdit(self)
        self.grid.addWidget(self.line, 1, 5)

        self.btn_update = QtWidgets.QPushButton("", self)
        self.btn_update.setIcon(QtGui.QIcon("update.png"))
        self.btn_update.clicked.connect(self.write)
        self.grid.addWidget(self.btn_update, 1, 6)

        self.lens = QtWidgets.QPushButton("", self)
        self.lens.setIcon(QtGui.QIcon("lens.png"))
        self.grid.addWidget(self.lens, 1, 7)

        self.btn_edit_record = QtWidgets.QPushButton("", self)
        self.btn_edit_record.setIcon(QtGui.QIcon("edit record.png"))
        self.btn_edit_record.clicked.connect(self.btn_edit_record_event)
        self.grid.addWidget(self.btn_edit_record, 1, 8)

        self.table = table

        self.grid.addWidget(self.table, 2, 1, 2, 30)
        self.setLayout(self.grid)

    def btn_edit_record_event(self):
        self.edit_record_1 = Class_commodity_cashiers_arm_window()
        self.edit_record_1.show()


class class_commodity_cashiers_arm(
    QtWidgets.QWidget, class_filtr, class_for_request_directory
):
    def __init__(self):
        super().__init__()

        self.btn_update = QtWidgets.QPushButton("Обновить все", self)
        self.btn_update.clicked.connect(self.write)

        self.to_excel = QtWidgets.QPushButton("В Excel все", self)
        self.to_excel.clicked.connect(self.write_excel)

        self.dict_class = {"АРМ товарного кассира": class_commodity_cashiers_arm_widget}

        self.dict_table = {"АРМ товарного кассира": None}

        self.grid = QtWidgets.QGridLayout()
        self.write_table()

        self.grid.addWidget(self.btn_update, 1, 1)
        self.grid.addWidget(self.to_excel, 1, 2)

        self.setLayout(self.grid)

    def write_table(self):
        for i in self.dict_table:
            self.dict_table[i] = QtWidgets.QTableWidget(self)
            self.dict_table[i].setColumnCount(0)
            self.dict_table[i].setRowCount(0)

            self.table_of_contents = self.dict_table[i].horizontalHeader()
            self.table_of_contents.sectionClicked.connect(self.click_table_of_contents)

            item = self.dict_class[i](self.dict_table[i])

            self.grid.addWidget(item, 2, 1, 1, 14)

    def write(self):
        self.dict_list = {
            "АРМ товарного кассира": list(
                csv.reader(open("АРМ товарного кассира.csv", "r"), delimiter=";")
            )
        }
        for i in self.dict_list:
            self.write_list(i, self.dict_list[i])

    def write_excel(self):
        os.chdir(
            str(
                QtWidgets.QFileDialog.getExistingDirectory(
                    self, "Выберите папку для сохранения"
                )
            )
        )
        wb = Workbook()

        for i in self.dict_list:
            wb.create_sheet(i)
            sheet = wb[i]
            for row in range(len(self.dict_list[i])):

                shiht_row = row + 1

                for column in range(len(self.dict_list[i][row])):
                    sheet.cell(row=shiht_row, column=column + 1).value = self.dict_list[
                        i
                    ][row][column]

            sheet.auto_filter.ref = "A1:AK100000"
            sheet.freeze_panes = sheet["A2"]
            sheet.sheet_view.zoomScale = 70
        wb.remove(wb["Sheet"])

        wb.save("excel.xlsx")
        print("excel.xlsx - done...")
